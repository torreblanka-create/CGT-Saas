import os
import sqlite3

import pandas as pd

from config.config import (
    COL_EXCEL_CATEGORIA,
    COL_EXCEL_DETALLE,
    COL_EXCEL_ID,
    COL_EXCEL_NOMBRE,
    DB_PATH,
    EXCEL_MAESTRO_PATH,
    EXCEL_SHEET_MAPPING,
)
from src.infrastructure.database import (
    ejecutar_query,
    get_db_connection,
    normalizar_texto,
    obtener_dataframe,
)


def obtener_ruta_excel():
    """Devuelve la ruta oficial del Excel definida en config.py"""
    return EXCEL_MAESTRO_PATH

def anexar_registro_maestro_excel(identificador, nombre, detalle, categoria, empresa, contrato):
    """
    Agrega de manera bidireccional un nuevo registro al Excel Físico (Base Maestra),
    buscando la última fila de la pestaña correspondiente sin alterar el formato.
    """
    import openpyxl
    ruta = obtener_ruta_excel()
    if not os.path.exists(ruta): return False

    hoja_objetivo = EXCEL_SHEET_MAPPING.get(categoria, "LISTADO_MAESTRO")

    try:
        wb = openpyxl.load_workbook(ruta)
        if hoja_objetivo not in wb.sheetnames:
            if "LISTADO_MAESTRO" in wb.sheetnames:
                hoja_objetivo = "LISTADO_MAESTRO"
            else:
                return False

        ws = wb[hoja_objetivo]

        # Encontrar la columna de cada campo basado en la primera fila (Headers)
        headers = {cell.value: idx for idx, cell in enumerate(ws[1], 1) if cell.value}

        nueva_fila = [None] * ws.max_column

        # Mapeo Inteligente (tolerante a espacios)
        for h_str, h_idx in headers.items():
            h_str_lower = str(h_str).strip().lower()
            if any(x in h_str_lower for x in ['identificador', 'rut', 'patente', 'id /']):
                nueva_fila[h_idx - 1] = identificador
            elif any(x in h_str_lower for x in ['nombre', 'descripcion', 'marca']):
                nueva_fila[h_idx - 1] = nombre
            elif 'detalle' in h_str_lower or 'cargo' in h_str_lower or 'modelo' in h_str_lower:
                nueva_fila[h_idx - 1] = detalle
            elif 'categor' in h_str_lower:
                nueva_fila[h_idx - 1] = categoria
            elif 'empresa' in h_str_lower:
                nueva_fila[h_idx - 1] = empresa
            elif 'contrato' in h_str_lower:
                nueva_fila[h_idx - 1] = contrato

        ws.append(nueva_fila)
        wb.save(ruta)
        wb.close()
        return True
    except Exception as e:
        print(f"Error escribiendo en Excel bidireccional: {e}")
        return False



def cargar_maestro_filtrado(categoria_ui, empresa_sel=None, contrato_sel=None):
    """
    Carga los datos maestros desde SQLite (maestro_entidades).
    Si no hay datos en la DB, intenta cargar desde Excel como fallback.
    """
    query = """
        SELECT 
            m.identificador as "Identificador",
            m.nombre as "Nombre",
            m.detalle as "Detalle",
            m.categoria as "Categoría",
            e.nombre as "Empresa",
            c.nombre_contrato as "Contrato"
        FROM maestro_entidades m
        LEFT JOIN empresas e ON m.empresa_id = e.id
        LEFT JOIN contratos c ON m.contrato_id = c.id
        WHERE m.categoria = ?
    """
    params = [categoria_ui]

    # Aplicar filtros de Empresa y Contrato si vienen definidos
    if empresa_sel and empresa_sel != "--- TODAS LAS EMPRESAS ---":
        query += " AND e.nombre = ?"
        params.append(empresa_sel)

    if contrato_sel and contrato_sel not in ["", "GLOBAL", "TODOS LOS CONTRATOS", "SIN CONTRATO"]:
        query += " AND c.nombre_contrato = ?"
        params.append(contrato_sel)

    try:
        df = obtener_dataframe(DB_PATH, query, tuple(params))

        # Fallback a Excel si la DB está vacía para esta categoría (opcional, para transición suave)
        if df.empty and os.path.exists(EXCEL_MAESTRO_PATH):
            df_excel = _cargar_desde_excel(categoria_ui, empresa_sel, contrato_sel)
            if not df_excel.empty:
                return df_excel

        return df

    except Exception as e:
        # La gestión de errores visuales debe hacerse en la capa de UI
        return pd.DataFrame()

def _cargar_desde_excel(categoria_ui, empresa_sel=None, contrato_sel=None):
    """Lógica original de carga desde Excel (ahora como fallback/importador)"""
    ruta_excel = EXCEL_MAESTRO_PATH
    if not os.path.exists(ruta_excel): return pd.DataFrame()

    nombre_hoja = EXCEL_SHEET_MAPPING.get(categoria_ui, "LISTADO_MAESTRO")
    try:
        df = pd.read_excel(ruta_excel, sheet_name=nombre_hoja, engine='openpyxl')
        df.columns = [str(c).strip() for c in df.columns]

        if empresa_sel and 'Empresa' in df.columns and empresa_sel != "--- TODAS LAS EMPRESAS ---":
            df = df[df['Empresa'].astype(str).str.strip().str.upper() == str(empresa_sel).strip().upper()]
        if contrato_sel and 'Contrato' in df.columns:
            if contrato_sel not in ["", "GLOBAL", "TODOS LOS CONTRATOS", "SIN CONTRATO"]:
                df = df[df['Contrato'].astype(str).str.strip().str.upper() == str(contrato_sel).strip().upper()]

        if COL_EXCEL_CATEGORIA in df.columns:
            df = df[df[COL_EXCEL_CATEGORIA].str.strip() == categoria_ui]

        return df
    except:
        return pd.DataFrame()


def obtener_listas_unicas(columna):
    """Extrae las empresas y contratos directamente de la DB"""
    try:
        if columna == "EMPRESA":
            query = "SELECT nombre FROM empresas ORDER BY nombre"
        else:
            query = "SELECT DISTINCT nombre_contrato FROM contratos ORDER BY nombre_contrato"

        df = obtener_dataframe(DB_PATH, query)
        return df.iloc[:, 0].tolist()
    except:
        return []


def obtener_contratos_por_empresa(empresa_sel):
    """Extrae SOLO los contratos que corresponden a la empresa seleccionada desde la DB"""
    try:
        query = """
            SELECT DISTINCT c.nombre_contrato 
            FROM contratos c
            JOIN empresas e ON c.empresa_id = e.id
            WHERE e.nombre = ?
            ORDER BY c.nombre_contrato
        """
        df = obtener_dataframe(DB_PATH, query, (empresa_sel,))
        return df['nombre_contrato'].tolist()
    except:
        return []

def exportar_maestro_a_excel():
    """Genera un DataFrame completo del maestro para exportar a Excel"""
    query = """
        SELECT 
            m.identificador as "Identificador",
            m.nombre as "Nombre",
            m.detalle as "Detalle",
            m.categoria as "Categoría",
            e.nombre as "Empresa",
            c.nombre_contrato as "Contrato"
        FROM maestro_entidades m
        LEFT JOIN empresas e ON m.empresa_id = e.id
        LEFT JOIN contratos c ON m.contrato_id = c.id
    """
    return obtener_dataframe(DB_PATH, query)

def sincronizar_maestro_desde_excel(emp_id, con_id, file_buffer=None):
    """Importa datos desde el Excel usando to_sql para mayor rendimiento."""
    try:
        # Se lee desde el buffer subido o desde la ruta por defecto
        if file_buffer is not None:
            df = pd.read_excel(file_buffer, engine='openpyxl')
        else:
            if not os.path.exists(EXCEL_MAESTRO_PATH):
                return False, f"No se encuentra el archivo Excel en {EXCEL_MAESTRO_PATH}"
            try:
                df = pd.read_excel(EXCEL_MAESTRO_PATH, sheet_name='LISTADO_MAESTRO', engine='openpyxl')
            except ValueError:
                df = pd.read_excel(EXCEL_MAESTRO_PATH, engine='openpyxl')

        df.columns = [str(c).strip() for c in df.columns]

        # Mapeo de columnas y limpieza
        df_db = pd.DataFrame()
        df_db['identificador'] = df.apply(lambda r: str(r.get(COL_EXCEL_ID, r.get('RUT', r.get('ID', '')))).strip(), axis=1)
        df_db['nombre'] = df.apply(lambda r: str(r.get(COL_EXCEL_NOMBRE, r.get('Nombre', r.get('Nombre_Completo', '')))).strip(), axis=1)
        df_db['detalle'] = df.apply(lambda r: str(r.get(COL_EXCEL_DETALLE, r.get('Cargo', r.get('Tipo', '')))).strip(), axis=1)
        df_db['categoria'] = df.apply(lambda r: str(r.get(COL_EXCEL_CATEGORIA, r.get('Categoria', 'Personal'))).strip(), axis=1)

        # ── NORMALIZACIÓN DE CATEGORÍAS (Trinity) ──────────────────────────────
        # Convierte nombres legacy/Excel al vocabulario interno de CGT.pro
        try:
            from config.config import CATEGORIA_ALIAS_MAP
            def _normalizar_categoria(cat_raw):
                c = str(cat_raw).strip()
                return CATEGORIA_ALIAS_MAP.get(c.lower(), c)
            df_db['categoria'] = df_db['categoria'].apply(_normalizar_categoria)
        except Exception:
            pass  # Si el alias map no está disponible, mantener valor original
        # ───────────────────────────────────────────────────────────────────────


        # Forzar IDs de empresa y contrato
        df_db['empresa_id'] = emp_id
        df_db['contrato_id'] = con_id

        # Eliminar filas vacías
        df_db = df_db[df_db['identificador'] != 'nan']
        df_db = df_db[df_db['identificador'] != '']

        count = len(df_db)
        if count == 0:
            return True, "No se encontraron registros válidos para sincronizar."

        with get_db_connection(DB_PATH) as conn:
            # Usar una tabla temporal para el "Upsert" masivo
            df_db.to_sql('temp_maestro', conn, if_exists='replace', index=False)

            # Ejecutar el UPSERT masivo
            conn.execute('''
                INSERT OR REPLACE INTO maestro_entidades (identificador, nombre, detalle, categoria, empresa_id, contrato_id)
                SELECT identificador, nombre, detalle, categoria, empresa_id, contrato_id FROM temp_maestro
            ''')
            conn.execute('DROP TABLE temp_maestro')
            conn.commit()

        return True, f"Se han sincronizado {count} registros exitosamente."
    except Exception as e:
        return False, f"Error en la sincronización: {str(e)}"

def actualizar_listado_maestro_sgi(empresa, contrato):
    """
    Genera y sobreescribe el archivo Excel de Control de Información Documentada
    según el estándar VELTV-F-SGI-QA-0001.
    """
    from datetime import datetime

    from src.infrastructure.archivos import obtener_ruta_modulo_especifico

    # 1. Obtener Datos
    query = """
        SELECT codigo, nombre, version, categoria, fecha_creacion, fecha_vencimiento
        FROM procedimientos
        WHERE (UPPER(empresa) = UPPER(?) OR empresa_id = (SELECT id FROM empresas WHERE UPPER(nombre) = UPPER(?) LIMIT 1))
    """
    params = [empresa, empresa]
    df = obtener_dataframe(DB_PATH, query, tuple(params))

    if df.empty:
        return False, "No hay documentos registrados para esta empresa."

    # 2. Lógica de Semáforo y Formateo
    hoy = datetime.now().date()

    def calcular_estado(row):
        f_creacion_str = row.get('fecha_creacion')
        if not f_creacion_str: return "S/I"
        try:
            f_creacion = datetime.strptime(f_creacion_str, '%Y-%m-%d').date()
            # Diferencia en meses (aprox)
            meses = (hoy.year - f_creacion.year) * 12 + (hoy.month - f_creacion.month)

            if meses >= 24: return "🔴 VENCIDO"
            if meses >= 23: return "🟡 POR VENCER (30 días)"
            return "🟢 VIGENTE"
        except:
            return "Error Fecha"

    df['Estado (Semaforo)'] = df.apply(calcular_estado, axis=1)

    # Renombrar columnas para el estándar solicitado
    filename = "260324 VELTV-F-SGI-QA-0001- Lista de Control de Información Documentada.xlsx"
    ruta_final = obtener_ruta_modulo_especifico(empresa, contrato, "SGI_DOCUMENTAL", crear=True)
    path_excel = os.path.join(ruta_final, filename)

    # Añadir nuevas columnas al Excel de salida
    columnas_sgi = {
        'codigo': 'Código Documento',
        'nombre': 'Nombre / Descripción',
        'version': 'Versión',
        'categoria': 'Categoría ISO',
        'fecha_creacion': 'Fecha Creación',
        'fecha_vencimiento': 'Próxima Revisión',
        'Estado (Semaforo)': 'Estatus Trazabilidad',
        'ambito': 'Ámbito',
        'sub_area': 'Sub-Área',
        'sigla_negocio': 'Sigla Negocio',
        'correlativo': 'Correlativo'
    }

    query = """
        SELECT codigo, nombre, version, categoria, fecha_creacion, fecha_vencimiento, 
               ambito, sub_area, sigla_negocio, correlativo
        FROM procedimientos
        WHERE (UPPER(empresa) = UPPER(?) OR empresa_id = (SELECT id FROM empresas WHERE UPPER(nombre) = UPPER(?) LIMIT 1))
    """
    df = obtener_dataframe(DB_PATH, query, (empresa, empresa))

    if not df.empty:
        df['Estado (Semaforo)'] = df.apply(calcular_estado, axis=1)
        df = df.rename(columns=columnas_sgi)
        # Ordenar columnas para el estándar
        cols_finales = [v for k, v in columnas_sgi.items() if v in df.columns]
        df = df[cols_finales]

    try:
        df.to_excel(path_excel, index=False, engine='openpyxl')
        return True, path_excel
    except Exception as e:
        return False, str(e)

def sincronizar_sgi_desde_excel(empresa_nom, contrato_nom, file_buffer=None):
    """
    Importa documentos SGI (PTS, Instructivos, etc.) desde un Excel 
    siguiendo el estándar VELTV-F-SGI-QA-0001 con detección robusta de cabeceras.
    """
    import os

    import pandas as pd

    from src.infrastructure.database import ejecutar_query, get_db_connection, normalizar_texto

    try:
        # 1. Leer el archivo (inicialmente sin cabeceras para buscar la fila de títulos)
        if file_buffer is not None:
            raw_df = pd.read_excel(file_buffer, header=None, engine='openpyxl')
        else:
            default_filename = "260324 VELTV-F-SGI-QA-0001- Lista de Control de Información Documentada.xlsx"
            if os.path.exists(default_filename):
                raw_df = pd.read_excel(default_filename, header=None, engine='openpyxl')
            else:
                return False, "No se proporcionó un archivo y no se encontró el archivo por defecto."

        # --- MOTOR DE DETECCIÓN DINÁMICA DE CABECERAS (MULTI-FILA) ---
        header_row_idx = -1
        keywords = ['Nombre del Documento', 'Código', 'Categoría ISO', 'Tipo de documento']

        # 1. Encontrar la primera fila que parece cabecera
        for i, row in raw_df.head(25).iterrows():
            row_str = " ".join([str(item) for item in row if pd.notna(item)])
            if any(key in row_str for key in keywords):
                header_row_idx = i
                break

        if header_row_idx == -1:
            return False, "No se encontró la fila de encabezados en el Excel (buscado en las primeras 25 filas)."

        # 2. Lógica de "Aplanamiento" de Cabecera (Soporta hasta 3 filas dinámicamente)
        header_rows = [raw_df.iloc[header_row_idx]]

        # Buscar si las siguientes 2 filas también son parte de la cabecera
        for offset in range(1, 3):
            next_idx = header_row_idx + offset
            if next_idx >= len(raw_df): break

            # Revisamos si la fila siguiente todavía parece cabecera (palabras clave)
            next_row = raw_df.iloc[next_idx]
            next_row_str = " ".join([str(item) for item in next_row if pd.notna(item)])
            if any(key in next_row_str for key in keywords):
                header_rows.append(next_row)
            else:
                break

        # Fusión Inteligente de Columnas (Aplanado)
        final_cols = []
        for col_idx in range(len(raw_df.columns)):
            parts = []
            for row in header_rows:
                val = str(row[col_idx]).strip() if pd.notna(row[col_idx]) else ""
                # Solo añadir si el valor es nuevo (De-duplicación vertical de celdas combinadas)
                if val and (not parts or val != parts[-1]):
                    parts.append(val)
            final_cols.append(" ".join(parts).strip())

        # Re-procesar el DataFrame usando la cabecera final combinada
        skip_rows = len(header_rows)
        df = raw_df.iloc[header_row_idx + skip_rows:].copy()
        df.columns = final_cols
        df = df.reset_index(drop=True)

        # 2. Mapeo de Columnas Ampliado (Estándar VELTV-F-SGI-QA-0001 Detallado + Simplificado)
        col_map = {
            # Formato Detallado (Fusión de Cabeceras)
            '2.- Estructura Código': 'codigo',
            '2.- Estructura Tipo de documento': 'categoria',
            '2.- Estructura Línea de Negocio (área)': 'sigla_negocio',
            '1.- Nombre del Documento': 'nombre',
            '5.- Revisión': 'version',
            '6.- Fecha creación o actualización': 'fecha_creacion',
            '7.- Fecha de Vencimiento': 'fecha_vencimiento',
            '8.- Estado del documento': 'estado_doc',
            'Estado': 'estado_doc',
            'Estatus': 'estado_doc',
            'Documento en si': 'path',
            'Ruta Final': 'path',
            'Ruta del Archivo': 'path',

            # Fallback a nombres simples
            'Código': 'codigo',
            'Nombre del Documento': 'nombre',
            'Tipo de documento': 'categoria',
            'Línea de Negocio (área)': 'sigla_negocio',
            'N°': 'correlativo',
            'Relator': 'asignado_a',

            # Formato Anterior / Otros
            'Código Documento': 'codigo',
            'Nombre / Descripción': 'nombre',
            'Versión': 'version',
            'Categoría ISO': 'categoria',
            'Fecha Creación': 'fecha_creacion',
            'Próxima Revisión': 'fecha_vencimiento',
            'Ámbito': 'ambito',
            'Sub-Área': 'sub_area',
            'Sigla Negocio': 'sigla_negocio',
            'Correlativo': 'correlativo',
            'Estado Documento': 'estado_doc'
        }

        # Adaptar nombres de columnas del DF al mapa interno
        found_cols = {}
        for col in df.columns:
            if col in col_map:
                found_cols[col_map[col]] = col

        # Validar columnas mínimas (Solo código y nombre son estrictamente obligatorios)
        if 'codigo' not in found_cols or 'nombre' not in found_cols:
            return False, f"Faltan columnas obligatorias: Se requiere 'Código' y 'Nombre'. Columnas detectadas: {list(df.columns[:5])}"

        # 3. Preparar DataFrame para SQLite
        df_db = pd.DataFrame()

        # Función auxiliar para extraer con el mapeo dinámico
        def get_val(row, internal_field, default=''):
            excel_col = found_cols.get(internal_field)
            if excel_col and excel_col in row:
                val = row[excel_col]
                return str(val).strip() if pd.notna(val) else default
            return default

        df_db['codigo'] = df.apply(lambda r: get_val(r, 'codigo').upper(), axis=1)
        df_db['nombre'] = df.apply(lambda r: get_val(r, 'nombre').upper(), axis=1)
        df_db['version'] = df.apply(lambda r: get_val(r, 'version', '0').upper(), axis=1)
        df_db['categoria'] = df.apply(lambda r: get_val(r, 'categoria', 'Documento'), axis=1)

        # Mapeo de nuevas columnas opcionales
        df_db['ambito'] = df.apply(lambda r: get_val(r, 'ambito'), axis=1)
        df_db['sub_area'] = df.apply(lambda r: get_val(r, 'sub_area'), axis=1)
        df_db['sigla_negocio'] = df.apply(lambda r: get_val(r, 'sigla_negocio').upper(), axis=1)
        df_db['correlativo'] = df.apply(lambda r: get_val(r, 'correlativo', ''), axis=1)

        # Fechas (intento de parseo flexible en motor SQLite posterior)
        df_db['fecha_creacion'] = df.apply(lambda r: get_val(r, 'fecha_creacion'), axis=1)
        df_db['fecha_vencimiento'] = df.apply(lambda r: get_val(r, 'fecha_vencimiento'), axis=1)
        df_db['estado_doc'] = df.apply(lambda r: get_val(r, 'estado_doc'), axis=1)

        # Lógica de Autodetención para Correlativo y Sigla si vienen vacíos pero el código existe
        def autoproteger_campos(row):
            cod = str(row['codigo']).strip()
            if not cod or cod == 'NAN': return row

            # Intentar reconocer patrones separando por guiones o puntos
            partes = cod.replace('.', '-').split('-')

            # Si la sigla de negocio está vacía, intentamos extraerla de la penúltima parte (ej: ...-QA-0001)
            if not row['sigla_negocio'] and len(partes) >= 2:
                # Usualmente la sigla de negocio es la penúltima o antepenúltima
                for p in reversed(partes[:-1]):
                    if len(p) == 2 or p.isalpha(): # Heurística: 2 letras suele ser la sigla de negocio
                        row['sigla_negocio'] = p.upper()
                        break

            # Si el correlativo está vacío, buscamos números en la última parte
            if not row['correlativo'] and len(partes) >= 1:
                import re
                nums = re.findall(r'\d+', partes[-1])
                if nums: row['correlativo'] = nums[0].zfill(4) # Intentar normalizar a 4 dígitos
                else:
                    # Buscar en todo el código si hay algún bloque puramente numérico
                    for p in reversed(partes):
                        if p.isdigit():
                            row['correlativo'] = p.zfill(4)
                            break
            return row

        df_db = df_db.apply(autoproteger_campos, axis=1)

        # Datos de contexto
        df_db['empresa'] = empresa_nom

        # Resolver IDs
        with get_db_connection(DB_PATH) as conn:
            cur = conn.cursor()
            cur.execute("SELECT id FROM empresas WHERE nombre = ?", (normalizar_texto(empresa_nom),))
            res_e = cur.fetchone()
            emp_id = res_e[0] if res_e else 0

            cur.execute("SELECT id FROM contratos WHERE empresa_id = ? AND nombre_contrato = ?", (emp_id, normalizar_texto(contrato_nom)))
            res_c = cur.fetchone()
            con_id = res_c[0] if res_c else 0

        df_db['empresa_id'] = emp_id
        df_db['contrato_id'] = con_id
        df_db['path'] = df.apply(lambda r: get_val(r, 'path', 'Sin archivo'), axis=1)

        # Eliminar filas sin código (vacías)
        df_db = df_db[df_db['codigo'] != 'NAN']
        df_db = df_db[df_db['codigo'] != '']

        count = len(df_db)
        if count == 0:
            return False, "No se encontraron registros válidos en el archivo."

        # 4. Inserción Atómica (UPSERT)
        with get_db_connection(DB_PATH) as conn:
            # PARCHE DE EMERGENCIA: Asegurar columnas críticas antes de la inserción
            for col_crit in ["path", "ambito", "sub_area", "sigla_negocio", "correlativo", "empresa", "estado_doc"]:
                try: conn.execute(f"ALTER TABLE procedimientos ADD COLUMN {col_crit} TEXT")
                except: pass

            df_db.to_sql('temp_sgi', conn, if_exists='replace', index=False)

            conn.execute('''
                INSERT INTO procedimientos (
                    codigo, nombre, version, fecha_creacion, fecha_vencimiento, 
                    categoria, empresa, empresa_id, contrato_id, path,
                    ambito, sub_area, sigla_negocio, correlativo
                )
                SELECT 
                    codigo, nombre, version, fecha_creacion, fecha_vencimiento, 
                    categoria, empresa, empresa_id, contrato_id, path,
                    ambito, sub_area, sigla_negocio, correlativo
                FROM temp_sgi
                WHERE 1
                ON CONFLICT(codigo) DO UPDATE SET
                    nombre=excluded.nombre,
                    version=excluded.version,
                    fecha_creacion=excluded.fecha_creacion,
                    fecha_vencimiento=excluded.fecha_vencimiento,
                    categoria=excluded.categoria,
                    empresa=excluded.empresa,
                    empresa_id=excluded.empresa_id,
                    contrato_id=excluded.contrato_id,
                    ambito=excluded.ambito,
                    sub_area=excluded.sub_area,
                    sigla_negocio=excluded.sigla_negocio,
                    correlativo=excluded.correlativo
            ''')
            conn.execute('DROP TABLE temp_sgi')
            conn.commit()

        # 5. SINCRONIZACIÓN INVERSA AUTOMÁTICA: Guardar copia en carpeta empresa
        actualizar_listado_maestro_sgi(empresa_nom, contrato_nom)

        return True, f"Se han sincronizado {count} documentos del SGI exitosamente."

    except Exception as e:
        return False, f"Error en sincronización SGI: {str(e)}"

